from fastapi import FastAPI, Depends, HTTPException, status, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import JSONResponse
from passlib.context import CryptContext
from starlette.middleware.sessions import SessionMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from fastapi import File, UploadFile
from fastapi import FastAPI, Query
import os
import shutil
import pandas as pd
from PIL import Image
import arabic_reshaper
from bidi.algorithm import get_display
from apscheduler.schedulers.background import BackgroundScheduler
import requests
from io import BytesIO
from pathlib import Path
from fastapi.staticfiles import StaticFiles
import openpyxl



# إعداد التطبيق
app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key="hkeinmnfo3e2ln")

# إعداد الباسورد
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# إعداد المجلدات
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

# تعريف المستخدم
class User(BaseModel):
    username: str
    password: str

def print_current_directory():
    current_directory = os.getcwd()  # الحصول على المسار الحالي
    print("المسار الحالي:", current_directory)

# دالة لفحص كلمة المرور
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

# دالة لتشفير كلمة المرور
def get_password_hash(password):
    return pwd_context.hash(password)
    
# دالة لتحويل النص العربي
def ar(text):
    reshaped_text = arabic_reshaper.reshape(text)
    bidi_text = get_display(reshaped_text)
    return bidi_text

# دالة لتفريغ محتويات المجلد وحذفه
def empty_and_delete_folder(folder_path):
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"فشل في حذف {file_path}: {e}")


# قاعدة بيانات المستخدمين (للتجربة فقط)
fake_users_db = {
    "ALAMIN_A": {
        "username": "ALAMIN_A",
        "hashed_password": get_password_hash("Asdfqwer1234@2")
    }
}

# التحقق من المستخدم
def get_user(db, username: str):
    if username in db:
        user_dict = db[username]
        return User(**user_dict)

# التحقق من المستخدم الحالي
def get_current_user(request: Request):
    username = request.session.get("username")
    if username in fake_users_db:
        return User(**fake_users_db[username])
    raise HTTPException(status_code=400, detail="Invalid credentials")


# الدالة التي ستنفذ كل خمس دقائق
def process_images():
    print("تشغيل العملية...")
    
    # استخدام الدالة لتفريغ محتويات المجلد وحذفه
    folder_path = 'static/uploads'
    empty_and_delete_folder(folder_path)
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    
    filexcl = BytesIO(response.content)
    
    # قراءة بيانات من ملف Excel
    df = pd.read_excel(filexcl, sheet_name='الصفحة الرئيسية', engine='openpyxl')

    # استخراج البيانات من الجدول
    q = df['الجاهز']
    w = df['اسم الموديل']
    p = len(w)

    # معالجة الصور وإضافة النص
    for i in range(p):
        if str(w[i]).isdigit():
            image_path = f"static/image/{w[i]}.jpg"
            if os.path.exists(image_path) and q[i] > 0:
                image = Image.open(image_path)
                
                # حفظ الصورة في المجلد
                image.save(f"static/uploads/رقم {w[i]} عدد الطقوم {q[i]}.jpg")

# دالة لتحديث الصور بناءً على عمود "الموجود"
def process_available_images():
    print("تشغيل عملية تحديث صور جديد...")
    
    # استخدام الدالة لتفريغ محتويات المجلد وحذفه
    folder_path = 'static/uploads'
    empty_and_delete_folder(folder_path)
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    
    filexcl = BytesIO(response.content)
    
    # قراءة بيانات من ملف Excel
    df = pd.read_excel(filexcl, sheet_name='الصفحة الرئيسية', engine='openpyxl')

    # استخراج البيانات من الجدول
    q = df['الموجود']
    w = df['اسم الموديل']
    p = len(w)

    # معالجة الصور وإضافة النص
    for i in range(p):
        if str(w[i]).isdigit():
            image_path = f"static/image/{w[i]}.jpg"
            if os.path.exists(image_path) and q[i] > 0:
                image = Image.open(image_path)
                
                # حفظ الصورة في المجلد
                image.save(f"static/uploads/رقم {w[i]} عدد الطقوم {q[i]}.jpg")


# دالة لتحديث الصور بناءً على عمود "النقص"
def process_available_images1():
    print("تشغيل عملية تحديث صور جديد...")
    
    # استخدام الدالة لتفريغ محتويات المجلد وحذفه
    folder_path = 'static/uploads'
    empty_and_delete_folder(folder_path)
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    
    filexcl = BytesIO(response.content)
    
    # قراءة بيانات من ملف Excel
    df = pd.read_excel(filexcl, sheet_name='الصفحة الرئيسية', engine='openpyxl')

    # استخراج البيانات من الجدول
    q = df['النقص']
    w = df['اسم الموديل']
    p = len(w)

    # معالجة الصور وإضافة النص
    for i in range(p):
        if str(w[i]).isdigit():
            image_path = f"static/image/{w[i]}.jpg"
            if os.path.exists(image_path) and q[i] > 0:
                image = Image.open(image_path)
                
                # حفظ الصورة في المجلد
                image.save(f"static/uploads/رقم {int(w[i])} عدد الطقوم {int(q[i])}.jpg")


# دالة لتحديث الصور بناءً على عمود "الأسعار"
def process_price_images():
    print("تشغيل عملية تحديث صور حسب الأسعار...")
    
    # استخدام الدالة لتفريغ محتويات المجلد وحذفه
    folder_path = 'static/uploads'
    empty_and_delete_folder(folder_path)
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    
    filexcl = BytesIO(response.content)
    
    # قراءة بيانات من ملف Excel
    df = pd.read_excel(filexcl, sheet_name='الصفحة الرئيسية', engine='openpyxl')

    # استخراج البيانات من الجدول
    q = df['السعر']
    w = df['اسم الموديل']
    p = len(w)

    # معالجة الصور وإضافة النص
    for i in range(p):
        if str(w[i]).isdigit():
            image_path = f"static/image/{w[i]}.jpg"
            if os.path.exists(image_path) and q[i] > 0:
                image = Image.open(image_path)
                
                # حفظ الصورة في المجلد
                image.save(f"static/uploads/رقم {w[i]} سعر {q[i]}.jpg")


# دالة لتحديث الصور بناءً على عمود "الأسعار لمحل ناظم
def process_price_nazm_images():
    print("تشغيل عملية تحديث صور حسب الأسعار...")
    
    # استخدام الدالة لتفريغ محتويات المجلد وحذفه
    folder_path = 'static/uploads'
    empty_and_delete_folder(folder_path)
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    
    filexcl = BytesIO(response.content)
    
    # قراءة بيانات من ملف Excel
    df = pd.read_excel(filexcl, sheet_name='الصفحة الرئيسية', engine='openpyxl')

    # استخراج البيانات من الجدول
    q = df['الأسعار لمحل ناظم']
    w = df['اسم الموديل']
    p = len(w)

    # معالجة الصور وإضافة النص
    for i in range(p):
        if str(w[i]).isdigit():
            image_path = f"static/image/{w[i]}.jpg"
            if os.path.exists(image_path) and q[i] > 0:
                image = Image.open(image_path)
                
                # حفظ الصورة في المجلد
                image.save(f"static/uploads/رقم {w[i]} السعر لمحل ناظم {q[i]}.jpg")



def process_pric_images():
    print("تشغيل عملية تحديث صور حسب رأس المال ...")
    
    # استخدام الدالة لتفريغ محتويات المجلد وحذفه
    folder_path = 'static/uploads'
    empty_and_delete_folder(folder_path)
    
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    
    filexcl = BytesIO(response.content)
    
    # قراءة بيانات من ملف Excel
    df = pd.read_excel(filexcl, sheet_name='بيانات رأس المال', engine='openpyxl')

    # استخراج البيانات من الجدول
    q = df['صافي التكلفة']
    w = df['الرقم']
    p = len(w)

    # معالجة الصور وإضافة النص
    for i in range(p):
        if str(w[i]).isdigit():
            image_path = f"static/image/{w[i]}.jpg"
            if os.path.exists(image_path) and q[i] > 0:
                image = Image.open(image_path)
                # إزالة الفواصل العشرية من رأس المال
                capital = int(q[i])
                # حفظ الصورة في المجلد مع النص المعدل
                image.save(f"static/uploads/رقم {w[i]} التكلفة {capital}.jpg")


# إعداد جدولة المهام
scheduler = BackgroundScheduler()
scheduler.add_job(process_images, 'interval', minutes=5)
scheduler.start()

# تحديد مسار مجلد templates
templates = Jinja2Templates(directory="templates")

# تحديد مسار مجلد لحفظ الصور المرفوعة
UPLOAD_DIR = Path("static/image")
UPLOAD_DIR.mkdir(exist_ok=True)

@app.post("/upload")
async def upload_image(file: UploadFile = File(...)):
    if file:
        file_location = UPLOAD_DIR / file.filename
        with open(file_location, "wb+") as file_object:
            file_object.write(file.file.read())
        return {"status": "success", "message": f"File '{file.filename}' uploaded successfully."}
    else:
        return {"status": "error", "message": "No file provided"}


UPLOAD_FOLDER = "static/image"

@app.delete("/delete_image")
async def delete_image(name: str = Query(...)):
    image_path = os.path.join(UPLOAD_FOLDER, f"{name}.jpg")
    if os.path.exists(image_path):
        os.remove(image_path)
        return {"success": True}
    else:
        return {"success": False, "error": "File not found"}


# الصفحة الرئيسية

@app.get("/", response_class=HTMLResponse)
async def read_index(request: Request):
    if "username" not in request.session:
        return RedirectResponse(url="/login")
    print_current_directory()
    images = [f.split('.')[0] for f in os.listdir('static/uploads') if os.path.isfile(os.path.join('static/uploads', f))]
    return templates.TemplateResponse("index.html", {"request": request, "images": images})

@app.get("/update-images")
async def update_images():
    try:
        process_images()
        return JSONResponse(content={"success": True})
    except Exception as e:
        print(f"حدث خطأ: {e}")
        return JSONResponse(content={"success": False})

@app.get("/update-available-images")
async def update_available_images():
    try:
        process_available_images()
        return JSONResponse(content={"success": True})
    except Exception as e:
        print(f"حدث خطأ: {e}")
        return JSONResponse(content={"success": False})

@app.get("/update-price-nazm-images1")
async def update_available_images2():
    try:
        process_available_images1()
        return JSONResponse(content={"success": True})
    except Exception as e:
        print(f"حدث خطأ: {e}")
        return JSONResponse(content={"success": False})

@app.get("/update-price-images")
async def update_price_images():
    try:
        process_price_images()
        return JSONResponse(content={"success": True})
    except Exception as e:
        print(f"حدث خطأ: {e}")
        return JSONResponse(content={"success": False})

@app.get("/update-pric-images") #التحديث بناءً على رأس المال
async def update_pric_images():
    try:
        process_pric_images()
        return JSONResponse(content={"success": True})
    except Exception as e:
        print(f"حدث خطأ: {e}")
        return JSONResponse(content={"success": False})

@app.get("/update-price-nazm-images") #التحديث بناء على الأسعار لمحل ناظم
async def update_price_nazm_images():
    try:
        process_price_nazm_images()
        return JSONResponse(content={"success": True})
    except Exception as e:
        print(f"حدث خطأ: {e}")
        return JSONResponse(content={"success": False})


# صفحة تسجيل الدخول
@app.get("/login", response_class=HTMLResponse)
async def login(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login", response_class=HTMLResponse)
async def login_for_access(request: Request, username: str = Form(...), password: str = Form(...)):
    user = fake_users_db.get(username)
    if user and verify_password(password, user['hashed_password']):
        request.session["username"] = username
        return RedirectResponse(url="/", status_code=status.HTTP_302_FOUND)
    raise HTTPException(status_code=400, detail="Invalid credentials")

# صفحة تسجيل الخروج
@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=status.HTTP_302_FOUND)

# صفحة الملخص
@app.get("/summary", response_class=HTMLResponse)
async def read_summary(request: Request):
    if "username" not in request.session:
        return RedirectResponse(url="/login")
    
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    filexcl = BytesIO(response.content)
    
    # قراءة بيانات من ملف Excel
    df = pd.read_excel(filexcl, sheet_name='ملخص', engine='openpyxl')
    
    # تحويل DataFrame إلى HTML
    summary_html = df.to_html(index=False, classes='summary-table')
    
    
    return templates.TemplateResponse("summary.html", {"request": request, "summary_html": summary_html})


@app.get("/summary1", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'ملخص البيع'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table1021':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table1021 غير موجود.")

    # تحديد نطاق الجدول
    min_row = int(table.ref.split(":")[0][1:])
    max_row = int(table.ref.split(":")[1][1:])
    
    # قراءة رؤوس الأعمدة
    headers = [cell.value for cell in ws[min_row]]
    
    # قراءة البيانات من الجدول (بدون رؤوس الأعمدة)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summar.html", {"request": request, "summary_html": data_html})


@app.get("/summary2", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'المصروف'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table15':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table15 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary1.html", {"request": request, "summary_html": data_html})


@app.get("/summary3", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'الزكاة'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table7':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table7 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = ref.split(":")[0][0], int(ref.split(":")[0][1:])
    max_col, max_row = ref.split(":")[1][0], int(ref.split(":")[1][1:])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary2.html", {"request": request, "summary_html": data_html})


@app.get("/summary4", response_class=HTMLResponse)
async def view_data(request: Request):
# رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'بيانات رأس المال'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'الجدول1':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول الجدول1 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)
    
    return templates.TemplateResponse("summary3.html", {"request": request, "summary_html": data_html})


@app.get("/summary5", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'البيانات'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table4':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table4 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary4.html", {"request": request, "summary_html": data_html})


@app.get("/summary6", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'الأمانات'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table19':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table19 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary5.html", {"request": request, "summary_html": data_html})


@app.get("/summary7", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'عبد العظيم'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table18':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table18 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary6.html", {"request": request, "summary_html": data_html})


@app.get("/summary8", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'عبد العظيم'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table17':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table17 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary7.html", {"request": request, "summary_html": data_html})


@app.get("/summary9", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'أحمد أحمدية'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'الجدول16':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول الجدول16 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary8.html", {"request": request, "summary_html": data_html})



@app.get("/summary10", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'الدفعات'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table3':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table3 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary9.html", {"request": request, "summary_html": data_html})


@app.get("/summary11", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'الإنتاج'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table2':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table2 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary10.html", {"request": request, "summary_html": data_html})


@app.get("/summary12", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'حساب خارجي'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table12':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table12 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary11.html", {"request": request, "summary_html": data_html})


@app.get("/summary13", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'صفحة الشباب الرئيسية'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table28':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table28 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary12.html", {"request": request, "summary_html": data_html})


@app.get("/summary14", response_class=HTMLResponse)
async def view_data(request: Request):
    # رابط التحميل المباشر للملف
    url = 'https://1drv.ms/x/c/84cafb26d1508441/EUGEUNEm-8oggITEBQAAAAABPBKjYUiBUdBKAh2dbwEYEw?download=1'
    
    # تحميل الملف من الرابط
    response = requests.get(url)
    file_path = BytesIO(response.content)
    sheet_name = 'الصفحة الرئيسية'  # استبدل هذا باسم الشيت الخاص بك

    # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # العثور على الجدول المحدد
    table = None
    for tbl in ws.tables.values():
        if tbl.name == 'Table11':
            table = tbl
            break
    if table is None:
        raise HTTPException(status_code=404, detail="الجدول Table11 غير موجود.")

    # تحليل نطاق الجدول
    ref = table.ref
    min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
    max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

    # تحويل أسماء الأعمدة إلى أرقام
    min_col_idx = openpyxl.utils.column_index_from_string(min_col)
    max_col_idx = openpyxl.utils.column_index_from_string(max_col)

    # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
    headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

    # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
    data = []
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx, values_only=True):
        data.append(row)

    # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
    df = pd.DataFrame(data, columns=headers)

    # عكس ترتيب الصفوف
    df = df.iloc[::-1].reset_index(drop=True)

    # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
    df = df.applymap(lambda x: round(x, 2) if isinstance(x, (int, float)) else x)

    # تحويل البيانات إلى HTML
    data_html = df.to_html(classes='summary-table', index=False)

    return templates.TemplateResponse("summary13.html", {"request": request, "summary_html": data_html})



@app.get("/upload_page", response_class=HTMLResponse)
async def upload_page(request: Request):
    if "username" not in request.session:
        return RedirectResponse(url="/login")
    with open("templates/upload.html", "r") as file:
        html_content = file.read()
    return HTMLResponse(content=html_content)

@app.get("/", response_class=HTMLResponse)
async def read_root():
    if "username" not in request.session:
        return RedirectResponse(url="/login")
    with open("templates/index.html", "r") as file:
        html_content = file.read()
    return HTMLResponse(content=html_content)

# صفحة الصور
@app.get("/images", response_class=HTMLResponse)
async def images(request: Request):
    if "username" not in request.session:
        return RedirectResponse(url="/login")
    
    images = sorted([int(f.split('.')[0]) for f in os.listdir('static/image') if os.path.isfile(os.path.join('static/image', f))], reverse=True)
    return templates.TemplateResponse("images.html", {"request": request, "images": images})
