from fastapi import HTTPException
from io import BytesIO
from PIL import Image
import pandas as pd
import json
import shutil
import requests
import main
import openpyxl
import os


class FanctionData():

    def load_json(self,filejson):
        with open(filejson, "r", encoding="utf-8") as file:
            return json.load(file)

    def update_title(self, NewTitle):
        data1 = FanctionData()
        data = data1.load_json(main.FileJSON)
        data["title"] = NewTitle

        with open(main.FileJSON, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=4)

    def empty_and_delete_folder(self,folder_path):
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print(f"فشل في حذف {file_path}: {e}")

    def update(self, col1, col2, sheet, imgname, url, NameTitle):

        #إعادة تسمية عنوان الصفحة
        updattitle = FanctionData()
        updattitle.update_title(NameTitle)

        # استخدام الدالة لتفريغ محتويات المجلد وحذفه
        folder_path = 'static/uploads'
        cleardata = FanctionData()
        cleardata.empty_and_delete_folder(folder_path)

        # تحميل الملف من الرابط
        response = requests.get(url)

        filexcl = BytesIO(response.content)

        # قراءة بيانات من ملف Excel
        df = pd.read_excel(filexcl, sheet_name=sheet, engine='openpyxl')

        # استخراج البيانات من الجدول
        q = df[col1]  # عمود سحب البيانات
        w = df[col2]  # عمود اسم البيانات
        p = len(w)

        # Rename Photos
        for i in range(p):
            if str(w[i]).isdigit():
                image_path = f"static/image/{w[i]}.jpg"
                if os.path.exists(image_path) and q[i] > 0:
                    image = Image.open(image_path)
                    captal = int(q[i])  # تحويل البيانات إلى رقم عشري
                    # حفظ الصورة في المجلد
                    image.save(f"static/uploads/رقم {w[i]} {imgname} {captal}.jpg")

    def viwe(self, url, sheet_name, name_table):
        # تحميل الملف من الرابط
        response = requests.get(url)
        file_path = BytesIO(response.content)

        # تحميل ملف إكسل لقراءة القيم المحسوبة فقط
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]

        # العثور على الجدول المحدد
        table = None
        for tbl in ws.tables.values():
            if tbl.name == name_table:
                table = tbl
                break
        if table is None:
            raise HTTPException(status_code=404, detail=f"الجدول {name_table} غير موجود.")

        # تحليل نطاق الجدول
        ref = table.ref
        min_col, min_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[0])
        max_col, max_row = openpyxl.utils.cell.coordinate_from_string(ref.split(":")[1])

        # تحويل أسماء الأعمدة إلى أرقام
        min_col_idx = openpyxl.utils.column_index_from_string(min_col)
        max_col_idx = openpyxl.utils.column_index_from_string(max_col)

        # قراءة رؤوس الأعمدة ضمن نطاق الجدول فقط
        headers = [ws.cell(row=min_row, column=col).value for col in range(min_col_idx, max_col_idx + 1)]

        # قراءة البيانات من الجدول فقط (ضمن نطاق الأعمدة والصفوف)
        data = []
        for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col_idx, max_col=max_col_idx,
                                values_only=True):
            data.append(row)

        # تحويل البيانات إلى DataFrame مع تحديد رؤوس الأعمدة
        df = pd.DataFrame(data, columns=headers)

        # عكس ترتيب الصفوف
        df = df.iloc[::-1].reset_index(drop=True)

        # تنسيق الأعمدة الرقمية لتحتوي على رقمين عشريين فقط
        df = df.apply(lambda col: col.map(lambda x: round(x, 2) if isinstance(x, (int, float)) else x))

        # تحويل البيانات إلى HTML
        data_html = df.to_html(classes='summary-table', index=False)
        return data_html